import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
import os

class registroProceso:

    def __init__(self, carpeta_salida, nombre_archivo_1, nombre_archivo_2):

        self.archivo_reporte_salida = os.path.join(carpeta_salida,nombre_archivo_1)
        self.escribir_encabezados()
        self.archivo_salida_ejecutivos = os.path.join(carpeta_salida,nombre_archivo_2)
        self.escribir_encabezados_ejecutivos()

    def escribir_encabezados(self):                 
        try:
            # Intenta cargar el archivo si ya existe
            workbook = openpyxl.load_workbook(self.archivo_reporte_salida)
        except FileNotFoundError:
            # Si el archivo no existe, lo crea
            workbook = Workbook()

        # Definimos los nombres de las hojas y encabezados
        hojas = ["Creacion", "Modificacion", "Activacion Desactivacion"]
        encabezados = ["Nombre Rol", "RUT", "Estado de Ejecución", "Error"]
        encabezados_cambios_de_roles = ["Nombre Rol", "RUT", "Estado de Ejecución", "Tipo de Modificación", "Error"]

        # Crear hojas y agregar encabezados si no existen ya
        for nombre_hoja in hojas:
            if nombre_hoja not in workbook.sheetnames:
                hoja = workbook.create_sheet(title=nombre_hoja)
                if nombre_hoja == "Creacion" or nombre_hoja == "Modificacion":
                    hoja.append(encabezados)  # Agrega los encabezados 
                else: 
                    hoja.append(encabezados_cambios_de_roles) #agrega los encabezados de cambios de roles


        # Borra la hoja predeterminada si está vacía y se ha creado al inicio
        if "Sheet" in workbook.sheetnames and not any(workbook["Sheet"].values):
            del workbook["Sheet"]

        workbook.save(self.archivo_reporte_salida)
        
    def escribir_resultado(self, tipo_operacion, nombre_rol, rut, estado_ejecucion, error):
        # Determina la hoja correspondiente según el tipo de operación
        if tipo_operacion.lower() == "creacion":
            nombre_hoja = "Creacion"
        elif tipo_operacion.lower() == "modificacion":
            nombre_hoja = "Modificacion"
        elif tipo_operacion.lower() == "activacion desactivacion":
            nombre_hoja = "Activacion Desactivacion"
        else:
            print(f"Operación desconocida: {tipo_operacion}")
            return
        # Cargar el archivo y seleccionar la hoja correspondiente
        workbook = load_workbook(self.archivo_reporte_salida)
        hoja = workbook[nombre_hoja]

        # Agregar una nueva fila con los datos
        nueva_fila = [nombre_rol, rut, estado_ejecucion, error]
        hoja.append(nueva_fila)

        # Guardar los cambios
        workbook.save(self.archivo_reporte_salida)
        workbook.close()

    def escribir_resultado_modificacion(self, tipo_operacion, nombre_rol, rut, estado_ejecucion, tipo_modificacion, error):
        # Determina la hoja correspondiente según el tipo de operación
        if tipo_operacion.lower() == "creacion":
            nombre_hoja = "Creacion"
        elif tipo_operacion.lower() == "modificacion":
            nombre_hoja = "Modificacion"
        elif tipo_operacion.lower() == "activacion desactivacion":
            nombre_hoja = "Activacion Desactivacion"
        else:
            print(f"Operación desconocida: {tipo_operacion}")
            return

        # Cargar el archivo y seleccionar la hoja correspondiente
        workbook = load_workbook(self.archivo_reporte_salida)
        hoja = workbook[nombre_hoja]

        # Agregar una nueva fila con los datos
        nueva_fila = [nombre_rol, rut, estado_ejecucion, tipo_modificacion, error]
        hoja.append(nueva_fila)

        # Guardar los cambios
        workbook.save(self.archivo_reporte_salida)
        workbook.close()

    def escribir_encabezados_ejecutivos(self):
        try:
            # Intenta cargar el archivo si ya existe
            workbook = openpyxl.load_workbook(self.archivo_salida_ejecutivos)
        except FileNotFoundError:
            # Si el archivo no existe, lo crea
            workbook = Workbook()

        # Definimos los nombres de las hojas y encabezados
        hojas = ["Altas", "Bajas","Carga Masiva de Roles"]
        encabezados_ejecutivos = ["Correo", "RUT", "Rol1", "Tipo Carga"]

        # Crear hojas y agregar encabezados si no existen ya
        for nombre_hoja in hojas:
            if nombre_hoja not in workbook.sheetnames:
                hoja = workbook.create_sheet(title=nombre_hoja)                
                hoja.append(encabezados_ejecutivos)  # Agrega los encabezados               


        # Borra la hoja predeterminada si está vacía y se ha creado al inicio
        if "Sheet" in workbook.sheetnames and not any(workbook["Sheet"].values):
            del workbook["Sheet"]

        workbook.save(self.archivo_salida_ejecutivos)

    def escribir_ejecutivos(self, correo, nombre_rol, rut, tipo_carga):
        # Determina la hoja correspondiente según el tipo de operación        
        nombre_hoja = "Carga Masiva de Roles"
        
        # Cargar el archivo y seleccionar la hoja correspondiente
        workbook = load_workbook(self.archivo_salida_ejecutivos)
        hoja = workbook[nombre_hoja]

        # Agregar una nueva fila con los datos
        nueva_fila = [correo, rut, nombre_rol, tipo_carga]
        hoja.append(nueva_fila)

        # Guardar los cambios
        workbook.save(self.archivo_salida_ejecutivos)
        workbook.close()
        
        